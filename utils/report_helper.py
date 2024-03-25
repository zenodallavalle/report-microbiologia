from itertools import combinations
from typing import Iterable, List, Optional
import openpyxl
import pandas as pd
import re

_widths = pd.read_csv("utils/widths.csv", index_col=[0, 1])


def add_overall_ws_intestation_row(ws):
    ws.cell(1, 1, "Periodo di riferimento")
    ws.cell(1, 2, "Tag materiale")
    ws.cell(1, 3, "Matrice")
    ws.cell(1, 4, "ID gruppo microbo")
    ws.cell(1, 5, "Gruppo microbo")
    ws.cell(1, 6, "Dettaglio isolati")
    ws.cell(1, 7, "Totale isolati")
    ws.cell(1, 8, "Resistenti (%)")
    ws.cell(1, 9, "Dettaglio resistenze")
    for cell in ws[1:1]:
        cell.font = openpyxl.styles.Font(bold=True)
        side = openpyxl.styles.Side(border_style="thin")
        cell.border = openpyxl.styles.Border(
            bottom=side, left=side, right=side, top=side
        )
        cell.alignment = openpyxl.styles.Alignment(horizontal="center")


def add_overall_rates_ws_intestation_row(ws):
    ws.cell(1, 1, "Periodo di riferimento")
    ws.cell(1, 2, "Tag materiale")
    ws.cell(1, 3, "Matrice")
    ws.cell(1, 4, "IDs gruppi microbi")
    ws.cell(1, 5, "Nome indicatore")
    ws.cell(1, 6, "Numeratore")
    ws.cell(1, 7, "Tasso per 10.000 gg")
    ws.cell(1, 8, "Tasso per 1.000 ricoveri o pazienti")
    ws.cell(1, 9, "Totale giornate di degenza")
    ws.cell(1, 10, "Totale ricoveri o pazienti")
    for cell in ws[1:1]:
        cell.font = openpyxl.styles.Font(bold=True)
        side = openpyxl.styles.Side(border_style="thin")
        cell.border = openpyxl.styles.Border(
            bottom=side, left=side, right=side, top=side
        )
        cell.alignment = openpyxl.styles.Alignment(horizontal="center")


def _get_width(column_letter: str, sheet_name: str = "_default"):
    try:
        return _widths.loc[(sheet_name, column_letter)].width
    except KeyError:
        return 10


def set_column_widths(wb):
    for ws in wb._sheets:
        sheet_name = ws.title
        width_name = sheet_name
        if sheet_name not in ("overall", "overall_rates"):
            width_name = "_default"
        for column in [c[0].column_letter for c in list(ws.columns)]:
            wb[sheet_name].column_dimensions[column].width = _get_width(
                column, sheet_name=width_name
            )


def set_wb_filters(wb):
    for ws in wb._sheets:
        if ws.title in ("overall"):
            if not len(ws["A"]):
                continue
            ws.auto_filter.ref = ws.dimensions
        elif ws.title in ("overall_rates"):
            if not len(ws["A"]) or not len(ws["K"]):
                continue
            left = ws["A"][0].coordinate
            right = ws["J"][-1].coordinate
            ws.auto_filter.ref = f"{left}:{right}"
        else:
            if not len(ws["A"]) or not len(ws["K"]):
                continue
            left = ws["A"][0].coordinate
            right = ws["K"][-1].coordinate
            ws.auto_filter.ref = f"{left}:{right}"


def generate_date_reference_for_excel_output(year: int, month: Optional[int] = None):
    if month:
        return pd.to_datetime(f"{year}-{month}-01").date()
    else:
        return f"{year}"


def simplify_resistances(resistances: Iterable):
    return list(set([re.sub(r">.*", "", resistance) for resistance in resistances]))


def _generate_resistenti_text_for_mdr_details(resistenze: pd.Series) -> List:
    unique_resistenze = resistenze.explode().unique()
    text_chunks = []
    for i in range(1, len(unique_resistenze) + 1):
        for resistances in combinations(unique_resistenze, i):
            mask = resistenze.apply(
                lambda resistenze_di_ogni_mo: all(
                    r in resistenze_di_ogni_mo for r in resistances
                )
            )
            n = mask.sum()
            if n == 0:
                continue
            legend = "DET -> " + "+".join(resistances).replace("MDR>", "").strip()
            text_chunks.append(f"{legend}: {n}")
    return text_chunks


def generate_resistenti_text_for_excel_output(df: pd.DataFrame):
    # fmt: off
    resistenze_no_details = df.resistente.replace("", pd.NA).dropna().str.replace('||', '|', regex=False).str.split("|").map(simplify_resistances, na_action='ignore')
    unique_resistenze = resistenze_no_details.explode().unique()
    mdr_details = df.resistente.replace("", pd.NA).dropna().str.split('|').explode()
    mdr_details = mdr_details[mdr_details.apply(lambda x: "MDR>" in x)]
    mdr_details = mdr_details.str.replace('MDR>', '', regex=False)
    # fmt: on
    text_chunks = []
    for i in range(1, len(unique_resistenze) + 1):
        for resistances in combinations(unique_resistenze, i):
            mask = resistenze_no_details.apply(
                lambda resistenze_di_ogni_mo: all(
                    r in resistenze_di_ogni_mo for r in resistances
                )
            )
            n = mask.sum()
            if n == 0:
                continue
            legend = "+".join(resistances)
            text_chunks.append(f"{legend}: {n}")
    resistenze = (
        df.resistente.replace("", pd.NA)
        .dropna()
        .str.replace("||", "|", regex=False)
        .str.split("|")
    )
    details_resistenze = _generate_resistenti_text_for_mdr_details(
        resistenze=resistenze
    )
    if len(details_resistenze) > 1:
        text_chunks += details_resistenze
    return "\n".join(text_chunks)
