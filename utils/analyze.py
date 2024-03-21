import openpyxl
import pandas as pd
from pandarallel import pandarallel
from .check_resistance import (
    check_resistance_and_validity,
    get_resistance_or_not_instructions,
)
from .duplicated_fns import (
    filter_query_repeated_isolation_in_patients_lt_1_month_resistance_wise,
    filter_query_repeated_isolation_in_patients_lt_1_month_no_resistance,
)
from .helper import (
    load_data,
    check_total_df,
    generate_excel_output_filename,
    filter_df_for_count,
)
from .report_helper import (
    add_overall_rates_ws_intestation_row,
    add_overall_ws_intestation_row,
    generate_date_reference_for_excel_output,
    generate_resistenti_text_for_excel_output,
    set_column_widths,
    set_wb_filters,
    simplify_resistances,
)

from .instructions import autogenerate

pandarallel.initialize(progress_bar=False, verbose=1)


def analyze(
    year,
    month,
    excel_output_folder_name="out",
    drop_column="to_drop",
    default_days_cutoff=30,
    instructions=list(),
    rate_instructions=list(),
    days_of_hospitalization=None,
    number_of_admissions_or_patients=None,
):
    total_df = load_data(year=year, month=month)
    check_total_df(total_df)

    resistance_instructions = get_resistance_or_not_instructions()

    # This df contains for each row the resistances of the microorganism isolated in that request. (One row = one microorganism)
    df = total_df.groupby(
        [
            "id_richiesta",
            "cognome_paziente",
            "nome_paziente",
            "data_nascita",
            "tags",
            "data_prelievo",
            "id_gruppo_microbo",
        ],
        as_index=False,
    ).parallel_apply(
        check_resistance_and_validity,
        keep_cols=[
            "nome_gruppo_microbo",
            "id_microbo",
            "nome_microbo",
            "id_reparto",
            "nome_reparto",
            "id_ricovero",
            "data_ricovero",
            "data_dimissione",
        ],
    )
    # fmt: off
    df["n_resistenze"] = (
        df.resistente.replace("", pd.NA)
        .dropna()
        .str.replace("||", "|", regex=False)  # The previous line will leave some || in the string
        .str.split("|")
        # .apply(simplify_resistances, na_action='ignore') # This delete the part that specifies for each MDR the nature of resistance
        # .map(lambda ress: len(set(ress)),na_action="ignore") # At this point if there was a MDR>xxx and MDR in the same row, they will be counted as 1
        .map(len, na_action="ignore") # This instead will simply count the number of resistances
    ) + (
        df.resistente.replace("", pd.NA)
        .dropna()
        .str.split("|")
        .map(lambda ress: any(["MDR" in res for res in ress]), na_action="ignore").replace(True, 0.1)  # Add a little bit of priority to MDR (.1)
    )
    # fmt: on
    # not_null_resistente_df contains only the rows with verfied resistance (if the antibiogram wasn't executed, the row is dropped)
    not_null_resistente_df = df.dropna(subset=["resistente"]).copy()

    excel_output_filepath = generate_excel_output_filename(
        excel_output_folder_name, year, month
    )

    # Do the actual analysis and write it to excel
    # data is an array of tuples (tag_materiale, nome_tag_materiale, id_gruppo_microbo, nome_gruppo_microbo, temp_df)
    data = []
    rates_data_no_mac_no_ps = []

    ## Add sorveglianza attiva to instructions for all available microorganisms
    autogeneration_mask = df.tags.str.split("|").apply(lambda tags: "sorv_att" in tags)
    autogeneration_mask &= df.data_prelievo.dt.year == year
    if month is not None:
        autogeneration_mask &= df.data_prelievo.dt.month == month

    _instructions = list(instructions)
    _instructions += autogenerate(
        df[autogeneration_mask], "sorv_att", "Sorveglianza attiva"
    )

    with pd.ExcelWriter(excel_output_filepath, engine="openpyxl") as writer:
        for instruction in _instructions:
            tag_materiale = instruction.tag
            nome_materiale = instruction.descrizione
            id_gruppo_microbo = instruction.gruppo_microbo_id
            nome_gruppo_microbo = instruction.descrizione_gruppo_microbo
            cutoff_repeat_days = instruction.cutoff_repeat_days
            resistenze_gruppo_microbo = resistance_instructions.get(
                id_gruppo_microbo, False
            )
            if resistenze_gruppo_microbo:
                temp_df = filter_df_for_count(
                    not_null_resistente_df,
                    year=year,
                    month=month,
                    resistances=None,
                    id_gruppo_microbo=id_gruppo_microbo,
                    tag=tag_materiale,
                    custom_filter_fn=filter_query_repeated_isolation_in_patients_lt_1_month_resistance_wise,
                    custom_filter_fn_kwargs=dict(
                        drop_column=drop_column,
                        days_cutoff=(
                            cutoff_repeat_days
                            if cutoff_repeat_days is not None
                            else default_days_cutoff
                        ),
                    ),
                )
                temp_df = temp_df.reindex(
                    columns=[
                        "id_richiesta",
                        "tag",
                        "tags",
                        "data_prelievo",
                        "nome_reparto",
                        "id_gruppo_microbo",
                        "nome_gruppo_microbo",
                        "id_microbo",
                        "nome_microbo",
                        "resistente",
                        "n_resistenze",
                        drop_column,
                    ]
                )
            else:
                temp_df = filter_df_for_count(
                    df,
                    year=year,
                    month=month,
                    id_gruppo_microbo=id_gruppo_microbo,
                    tag=tag_materiale,
                    custom_filter_fn=filter_query_repeated_isolation_in_patients_lt_1_month_no_resistance,
                    custom_filter_fn_kwargs=dict(
                        drop_column=drop_column,
                        days_cutoff=(
                            cutoff_repeat_days
                            if cutoff_repeat_days is not None
                            else default_days_cutoff
                        ),
                    ),
                )
                temp_df = temp_df.reindex(
                    columns=[
                        "id_richiesta",
                        "tag",
                        "tags",
                        "data_prelievo",
                        "nome_reparto",
                        "id_gruppo_microbo",
                        "nome_gruppo_microbo",
                        "id_microbo",
                        "nome_microbo",
                        drop_column,
                    ]
                )

            # convert datetime to date
            if "data_prelievo" in temp_df.columns:
                temp_df["data_prelievo"] = temp_df.data_prelievo.dt.date
            if "data_ricovero" in temp_df.columns:
                temp_df["data_ricovero"] = temp_df.data_ricovero.dt.date
            if "data_dimissione" in temp_df.columns:
                temp_df["data_dimissione"] = temp_df.data_dimissione.dt.date
            sheet_name = f"{tag_materiale}-{id_gruppo_microbo}"
            data.append(
                (
                    tag_materiale,
                    nome_materiale,
                    id_gruppo_microbo,
                    nome_gruppo_microbo,
                    temp_df,
                )
            )
            temp_df.to_excel(writer, sheet_name=sheet_name, index=False)

    # No mac no ps for rates
    df = df[
        (~df.nome_reparto.str.contains("mac", regex=False, case=False, na=False))
        & (
            ~df.nome_reparto.str.contains(
                "pronto soccorso", regex=False, case=False, na=False
            )
        )
        & (~df.id_ricovero.str.startswith("PS", na=False))
    ]
    df["resistente"] = df.resistente.fillna("")

    for instruction in rate_instructions:
        tag_materiale = instruction.tag
        nome_materiale = instruction.descrizione
        id_gruppo_microbi = (
            instruction.id_gruppo_microbi
            if not isinstance(instruction.id_gruppo_microbi, str)
            else [instruction.id_gruppo_microbi]
        )
        nome = instruction.descrizione_gruppo_microbi
        cutoff_repeat_days = instruction.cutoff_repeat_days
        select_fn = instruction.select_fn
        temp_dfs = []
        for id_gruppo_microbo in id_gruppo_microbi:
            resistenze_gruppo_microbo = resistance_instructions.get(
                id_gruppo_microbo, False
            )
            if resistenze_gruppo_microbo:
                temp_dfs.append(
                    filter_df_for_count(
                        df,
                        year=year,
                        month=month,
                        resistances=None,
                        id_gruppo_microbo=id_gruppo_microbo,
                        tag=tag_materiale,
                        custom_filter_fn=filter_query_repeated_isolation_in_patients_lt_1_month_resistance_wise,
                        custom_filter_fn_kwargs=dict(
                            drop_column=drop_column,
                            days_cutoff=(
                                cutoff_repeat_days
                                if cutoff_repeat_days is not None
                                else default_days_cutoff
                            ),
                        ),
                    )
                )
            else:
                temp_dfs.append(
                    filter_df_for_count(
                        df,
                        year=year,
                        month=month,
                        id_gruppo_microbo=id_gruppo_microbo,
                        tag=tag_materiale,
                        custom_filter_fn=filter_query_repeated_isolation_in_patients_lt_1_month_no_resistance,
                        custom_filter_fn_kwargs=dict(
                            drop_column=drop_column,
                            days_cutoff=(
                                cutoff_repeat_days
                                if cutoff_repeat_days is not None
                                else default_days_cutoff
                            ),
                        ),
                    )
                )
        temp_df = pd.concat(temp_dfs).reindex(
            columns=[
                "id_richiesta",
                "tag",
                "tags",
                "data_prelievo",
                "nome_reparto",
                "id_gruppo_microbo",
                "nome_gruppo_microbo",
                "id_microbo",
                "nome_microbo",
                "resistente",
                "n_resistenze",
                drop_column,
            ]
        )
        # convert datetime to date
        if "data_prelievo" in temp_df.columns:
            temp_df["data_prelievo"] = temp_df.data_prelievo.dt.date
        if "data_ricovero" in temp_df.columns:
            temp_df["data_ricovero"] = temp_df.data_ricovero.dt.date
        if "data_dimissione" in temp_df.columns:
            temp_df["data_dimissione"] = temp_df.data_dimissione.dt.date
        temp_df["is_numerator"] = select_fn(temp_df)
        rates_data_no_mac_no_ps.append(
            (
                tag_materiale,
                nome_materiale,
                ", ".join(id_gruppo_microbi),
                nome,
                temp_df,
            )
        )

    wb = openpyxl.load_workbook(excel_output_filepath)
    sheetnames = list(wb.sheetnames)

    # Add overall sheet and hyperlinks
    overall_ws = wb.create_sheet("overall")
    overall_rates_ws = wb.create_sheet("overall_rates")

    # Intestation
    add_overall_ws_intestation_row(overall_ws)
    add_overall_rates_ws_intestation_row(overall_rates_ws)

    # Custom formats
    reference_date_format = openpyxl.styles.NamedStyle(
        name="cd2", number_format="MMM-YY"
    )

    # Overall ws: iterate data
    for i, (
        tag_materiale,
        nome_tag_materiale,
        id_gruppo_microbo,
        nome_gruppo_microbo,
        temp_df,
    ) in enumerate(data, start=2):
        temp_df = temp_df[temp_df[drop_column].isnull()]
        try:
            if temp_df.resistente.dropna().empty:
                resistenti = None
            else:
                resistenti = temp_df.resistente.astype(bool).sum()
        except (AttributeError, KeyError):
            resistenti = None
        sheet_name = f"{tag_materiale}-{id_gruppo_microbo}"
        tot = len(temp_df)

        overall_ws.cell(i, 1, generate_date_reference_for_excel_output(year, month))
        overall_ws.cell(i, 1).style = reference_date_format
        overall_ws.cell(i, 2, tag_materiale)
        overall_ws.cell(i, 3, nome_tag_materiale)
        overall_ws.cell(i, 4, id_gruppo_microbo)
        overall_ws.cell(i, 5, nome_gruppo_microbo)

        overall_ws.cell(i, 6).hyperlink = f"#'{sheet_name}'!A1"
        overall_ws.cell(i, 6).value = "vai al dettaglio"
        overall_ws.cell(i, 6).style = "Hyperlink"

        if tot is not None:
            # Add tot
            overall_ws.cell(i, 7, tot)
            if resistenti is not None:
                # Add n resistenti and pct
                overall_ws.cell(
                    i, 8, f"{resistenti} ({resistenti/tot:.0%})".replace(".", ",")
                )
                # add dettaglio resistenze if any resistenti
                if resistenti:
                    resistenti_text = generate_resistenti_text_for_excel_output(temp_df)
                    overall_ws.cell(i, 9, resistenti_text)
                    overall_ws.cell(i, 9).alignment = openpyxl.styles.Alignment(
                        wrap_text=True
                    )

        # Add link to overall sheet
        sheet = wb[sheet_name]
        sheet["L1"].hyperlink = f"#'overall'!A{i}"
        sheet["L1"].value = "torna all'indice"
        sheet["L1"].style = "Hyperlink"

    # Overall_rates_ws: iterate rates_data
    for i, (
        tag_materiale,
        nome_materiale,
        id_gruppo_microbi,
        nome,
        temp_df,
    ) in enumerate(rates_data_no_mac_no_ps, start=2):
        temp_df = temp_df[temp_df[drop_column].isnull()]
        try:
            numeratore = temp_df.is_numerator.sum()
            if pd.isnull(numeratore):
                numeratore = 0
        except (AttributeError, KeyError):
            numeratore = None

        overall_rates_ws.cell(
            i, 1, generate_date_reference_for_excel_output(year, month)
        )
        overall_rates_ws.cell(i, 1).style = reference_date_format
        overall_rates_ws.cell(i, 2, tag_materiale)
        overall_rates_ws.cell(i, 3, nome_materiale)
        overall_rates_ws.cell(i, 4, id_gruppo_microbi)
        overall_rates_ws.cell(i, 4).alignment = openpyxl.styles.Alignment(
            wrap_text=True
        )
        overall_rates_ws.cell(i, 5, nome)
        if days_of_hospitalization is not None:
            overall_rates_ws.cell(i, 9, days_of_hospitalization)
        if number_of_admissions_or_patients is not None:
            overall_rates_ws.cell(i, 10, number_of_admissions_or_patients)
        if numeratore is not None:
            overall_rates_ws.cell(i, 6, numeratore)
            if days_of_hospitalization is not None:
                # Add tasso per 10.000 gg as formula
                overall_rates_ws.cell(i, 7, f"=F{i}/I{i}*10000")
                overall_rates_ws.cell(i, 7).number_format = "0.00"
            if number_of_admissions_or_patients is not None:
                # Add tasso per 1.000 ricoveri or pazienti as formula
                overall_rates_ws.cell(i, 8, f"=F{i}/J{i}*1000")
                overall_rates_ws.cell(i, 8).number_format = "0.00"
    # Add link to overall overall_rates_ws
    overall_rates_ws["L1"].hyperlink = "#'overall'!A1"
    overall_rates_ws["L1"].value = "torna all'indice"
    overall_rates_ws["L1"].style = "Hyperlink"

    # Set K column .styles.Alignment(wrap_text=True) for all sheets except overall
    for ws in wb._sheets:
        if ws.title not in ("overall", "overall_rates"):
            for cell in ws["K"]:
                cell.alignment = openpyxl.styles.Alignment(wrap_text=True)

    # Set column width
    set_column_widths(wb)

    # Set filters
    set_wb_filters(wb)

    # Set sheet order
    wb._sheets = [
        wb[sheetname] for sheetname in ["overall", "overall_rates"] + sheetnames
    ]

    # Save
    wb.save(excel_output_filepath)
